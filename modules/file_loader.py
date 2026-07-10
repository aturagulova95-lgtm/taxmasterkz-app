"""
file_loader.py
Универсальная загрузка файлов Excel/CSV/PDF с автоопределением строки заголовка,
листов книги и базовой очисткой "мусорных" верхних/пустых строк.

Задача модуля — превратить произвольный файл в список "сырых" таблиц
(pandas.DataFrame), готовых к передаче в column_mapper / data_normalizer.
Модуль не знает про налоговую специфику — это чисто техническая загрузка.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None

try:
    import pdfplumber
except ImportError:  # pragma: no cover
    pdfplumber = None


# Ключевые слова, по которым мы ищем "настоящую" строку заголовка среди
# первых N строк листа (часто выгрузки КГД содержат 1-3 служебные строки
# сверху: название отчета, дата выгрузки и т.д.)
HEADER_KEYWORDS = [
    "дата", "бин", "иин", "сумма", "ндс", "номер", "рег", "наименование",
    "статус", "код", "период", "квартал", "вид", "стоимость", "оборот",
    "количество", "цена", "снт", "год", "инспектор",
]

MAX_HEADER_SCAN_ROWS = 15


@dataclass
class LoadedTable:
    """Одна "сырая" таблица, извлечённая из файла (лист Excel или CSV)."""

    source_file: str
    sheet_name: str
    header_row_index: int  # 0-based индекс строки-заголовка в исходном листе
    dataframe: pd.DataFrame
    warnings: list[str] = field(default_factory=list)


@dataclass
class LoadedPdf:
    source_file: str
    text: str
    pages: list[str]


def _score_header_row(cells: list) -> int:
    """Оценивает, насколько строка похожа на заголовок таблицы."""
    score = 0
    non_empty = 0
    for c in cells:
        if c is None:
            continue
        s = str(c).strip()
        if not s:
            continue
        non_empty += 1
        low = s.lower()
        if any(kw in low for kw in HEADER_KEYWORDS):
            score += 2
        # заголовки обычно текстовые и не слишком длинные
        if len(s) < 60 and not re.match(r"^-?\d+([.,]\d+)?$", s):
            score += 1
    if non_empty < 2:
        return 0
    return score


def detect_header_row(rows: list[list]) -> int:
    """
    Находит индекс (0-based) наиболее вероятной строки заголовка
    среди первых MAX_HEADER_SCAN_ROWS строк.
    Если ни одна строка не набрала достаточный балл — возвращает 0.
    """
    best_idx, best_score = 0, -1
    for i, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
        score = _score_header_row(row)
        # бонус, если следующая строка выглядит как строка данных
        # (в ней есть числа/даты), а сама строка — почти вся текст
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx if best_score > 0 else 0


def _dedupe_columns(cols: list[str]) -> list[str]:
    seen: dict[str, int] = {}
    out = []
    for c in cols:
        c = (c or "").strip() or "col"
        if c in seen:
            seen[c] += 1
            out.append(f"{c}__{seen[c]}")
        else:
            seen[c] = 0
            out.append(c)
    return out


def load_excel(file_bytes: bytes, filename: str) -> list[LoadedTable]:
    """
    Читает все листы Excel-файла (xlsx/xls), для каждого листа
    автоматически находит строку заголовка и возвращает DataFrame
    с "чистыми" данными (без служебных строк сверху и пустых строк).
    """
    results: list[LoadedTable] = []
    bio = io.BytesIO(file_bytes)

    try:
        wb = openpyxl.load_workbook(bio, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
    except Exception:
        # fallback: pandas сам разберётся (например, старый .xls формат)
        try:
            xls = pd.ExcelFile(io.BytesIO(file_bytes))
            sheet_names = xls.sheet_names
        except Exception as e:
            raise ValueError(f"Не удалось открыть Excel-файл {filename}: {e}")
        wb = None

    if wb is not None:
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                rows.append(list(row))
                if i > 5000 and len(rows) > 5000:
                    # защитная граница для сканирования заголовка,
                    # полный лист всё равно читаем ниже через pandas
                    pass
            if not rows:
                continue
            header_idx = detect_header_row(rows)
            header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
            header = _dedupe_columns(header)
            data_rows = rows[header_idx + 1 :]
            df = pd.DataFrame(data_rows, columns=header)
            # убираем полностью пустые строки/столбцы
            df = df.dropna(axis=0, how="all")
            df = df.dropna(axis=1, how="all")
            df = df.loc[:, [c for c in df.columns if c and not c.startswith("col")]] \
                if any(c and not c.startswith("col") for c in df.columns) else df
            warnings = []
            if header_idx > 0:
                warnings.append(
                    f"Заголовок найден в строке {header_idx + 1} "
                    f"(строки 1-{header_idx} пропущены как служебные)"
                )
            results.append(
                LoadedTable(
                    source_file=filename,
                    sheet_name=sheet_name,
                    header_row_index=header_idx,
                    dataframe=df.reset_index(drop=True),
                    warnings=warnings,
                )
            )
        wb.close()
    else:
        for sheet_name in sheet_names:
            raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None)
            rows = raw.values.tolist()
            header_idx = detect_header_row(rows)
            header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
            header = _dedupe_columns(header)
            df = pd.DataFrame(rows[header_idx + 1 :], columns=header)
            df = df.dropna(axis=0, how="all")
            results.append(
                LoadedTable(
                    source_file=filename,
                    sheet_name=sheet_name,
                    header_row_index=header_idx,
                    dataframe=df.reset_index(drop=True),
                )
            )
    return results


def load_csv(file_bytes: bytes, filename: str) -> list[LoadedTable]:
    """Читает CSV с автоопределением разделителя и строки заголовка."""
    text = None
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = file_bytes.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = file_bytes.decode("utf-8", errors="replace")

    lines = text.splitlines()
    sample_rows = [re.split(r"[;,\t]", line) for line in lines[:MAX_HEADER_SCAN_ROWS]]
    header_idx = detect_header_row(sample_rows)

    sep = ";" if lines and lines[0].count(";") >= lines[0].count(",") else ","
    try:
        df = pd.read_csv(io.StringIO(text), sep=sep, skiprows=header_idx, engine="python")
    except Exception:
        df = pd.read_csv(io.StringIO(text), sep=None, skiprows=header_idx, engine="python")
    df = df.dropna(axis=0, how="all")
    warnings = []
    if header_idx > 0:
        warnings.append(f"Заголовок найден в строке {header_idx + 1}")
    return [
        LoadedTable(
            source_file=filename,
            sheet_name="csv",
            header_row_index=header_idx,
            dataframe=df.reset_index(drop=True),
            warnings=warnings,
        )
    ]


def load_pdf(file_bytes: bytes, filename: str) -> LoadedPdf:
    """Извлекает текст из PDF постранично."""
    if pdfplumber is None:
        raise RuntimeError("pdfplumber не установлен")
    pages = []
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for p in pdf.pages:
            pages.append(p.extract_text() or "")
    return LoadedPdf(source_file=filename, text="\n".join(pages), pages=pages)


def load_any(file_bytes: bytes, filename: str):
    """
    Единая точка входа: определяет тип файла по расширению
    и вызывает соответствующий загрузчик.
    Возвращает либо list[LoadedTable] (для excel/csv), либо LoadedPdf.
    """
    lower = filename.lower()
    if lower.endswith((".xlsx", ".xlsm", ".xls")):
        return load_excel(file_bytes, filename)
    if lower.endswith(".csv"):
        return load_csv(file_bytes, filename)
    if lower.endswith(".pdf"):
        return load_pdf(file_bytes, filename)
    raise ValueError(f"Неподдерживаемый формат файла: {filename}")
