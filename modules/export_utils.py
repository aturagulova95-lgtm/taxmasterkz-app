"""
export_utils.py
Экспорт аналитических таблиц в Excel: свод по покупателям, свод по
поставщикам, свод по расхождениям (риски) — с цветовой индикацией
(зелёный/жёлтый/красный) по уровню риска/расхождения.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2C5F8A", end_color="2C5F8A", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)

LEVEL_FILL = {"высокий": RED, "средний": YELLOW, "низкий": GREEN}


def _autosize(ws, df: pd.DataFrame):
    for i, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str).tolist()[:200]])
        ws.column_dimensions[get_column_letter(i)].width = min(max(12, max_len + 2), 60)


def _write_sheet(writer, df: pd.DataFrame, sheet_name: str, level_col: str | None = None):
    if df.empty:
        df = pd.DataFrame([{"Информация": "Данные отсутствуют"}])
    df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    ws = writer.sheets[sheet_name[:31]]
    for i, _ in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    if level_col and level_col in df.columns:
        col_idx = list(df.columns).index(level_col) + 1
        for row in range(2, len(df) + 2):
            val = ws.cell(row=row, column=col_idx).value
            fill = LEVEL_FILL.get(str(val).lower() if val else "", None)
            if fill:
                for c in range(1, len(df.columns) + 1):
                    ws.cell(row=row, column=c).fill = fill
    _autosize(ws, df)
    ws.freeze_panes = "A2"


def export_buyers_excel(buyers_df: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        _write_sheet(writer, buyers_df, "Покупатели")
    return bio.getvalue()


def export_suppliers_excel(suppliers_df: pd.DataFrame) -> bytes:
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        _write_sheet(writer, suppliers_df, "Поставщики")
    return bio.getvalue()


def export_findings_excel(findings: list[dict]) -> bytes:
    df = pd.DataFrame(findings)
    if not df.empty:
        cols = ["level", "risk_type", "period", "amount", "counterparty_name", "counterparty_bin",
                "description", "what_to_check", "possible_consequence", "source", "source_ref"]
        df = df[[c for c in cols if c in df.columns]]
        df = df.rename(columns={
            "level": "Уровень риска", "risk_type": "Вид риска", "period": "Период", "amount": "Сумма",
            "counterparty_name": "Контрагент", "counterparty_bin": "БИН контрагента",
            "description": "Описание", "what_to_check": "Что проверить",
            "possible_consequence": "Возможные последствия", "source": "Источник", "source_ref": "Ссылка на данные",
        })
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        _write_sheet(writer, df, "Риски", level_col="Уровень риска")
    return bio.getvalue()


def export_basic_workbook(dynamics_df: pd.DataFrame, buyers_df: pd.DataFrame,
                           suppliers_df: pd.DataFrame) -> bytes:
    """Урезанная книга только с «основными сводами» (динамика, покупатели,
    поставщики) — для тарифа START, где расширенные выгрузки (риски, сверка
    с ФНО) закрыты (см. ТЗ v4 п.9)."""
    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        _write_sheet(writer, dynamics_df, "Динамика по годам")
        _write_sheet(writer, buyers_df, "Покупатели")
        _write_sheet(writer, suppliers_df, "Поставщики")
    return bio.getvalue()


def export_full_workbook(dynamics_df: pd.DataFrame, buyers_df: pd.DataFrame,
                          suppliers_df: pd.DataFrame, findings: list[dict],
                          vat_rec_df: pd.DataFrame, income_rec_df: pd.DataFrame) -> bytes:
    """Единая книга со всеми сводами — приложение к аналитической справке."""
    findings_df = pd.DataFrame(findings)
    if not findings_df.empty:
        findings_df = findings_df.rename(columns={
            "level": "Уровень риска", "risk_type": "Вид риска", "period": "Период", "amount": "Сумма",
            "counterparty_name": "Контрагент", "counterparty_bin": "БИН контрагента",
            "description": "Описание", "what_to_check": "Что проверить",
            "possible_consequence": "Возможные последствия",
        })
        keep = ["Уровень риска", "Вид риска", "Период", "Сумма", "Контрагент", "БИН контрагента",
                "Описание", "Что проверить", "Возможные последствия"]
        findings_df = findings_df[[c for c in keep if c in findings_df.columns]]

    bio = io.BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        _write_sheet(writer, dynamics_df, "Динамика по годам")
        _write_sheet(writer, buyers_df, "Покупатели")
        _write_sheet(writer, suppliers_df, "Поставщики")
        _write_sheet(writer, vat_rec_df, "НДС ЭСФ vs ФНО 300")
        _write_sheet(writer, income_rec_df, "Доход ЭСФ vs ФНО 100")
        _write_sheet(writer, findings_df, "Риски", level_col="Уровень риска")
    return bio.getvalue()
